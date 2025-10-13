#!/usr/bin/env python3
"""
Final Comprehensive IT Helpdesk Report Enhancement
Creates industry-standard reporting with visuals, analytics, and insights
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from datetime import datetime
import re

# File paths
enhanced_file = "/Users/matt/Library/CloudStorage/OneDrive-Ever.Ag(2)/Desktop/Monthly_Reporting_Aug_Enhanced.xlsx"
original_file = "/Users/matt/Library/CloudStorage/OneDrive-Ever.Ag(2)/Desktop/Monthly_Reporting_Aug.xlsx"

def analyze_and_enhance_data():
    """Comprehensive analysis and enhancement of helpdesk data"""
    
    print("🔍 Loading and analyzing helpdesk data...")
    
    # Load data
    time_df = pd.read_excel(original_file, sheet_name='Time Spent')
    help_ticket_df = pd.read_excel(original_file, sheet_name='Help Ticket')
    
    # Clean and process time data
    def time_to_hours(time_str):
        if pd.isna(time_str) or time_str == 'Not Found':
            return 0
        try:
            time_str = str(time_str).strip()
            hours = 0
            minutes = 0
            
            if 'h' in time_str and 'm' in time_str:
                parts = time_str.split('h')
                hours = int(parts[0].strip())
                minutes = int(parts[1].replace('m', '').strip())
            elif 'h' in time_str:
                hours = int(time_str.replace('h', '').strip())
            elif 'm' in time_str:
                minutes = int(time_str.replace('m', '').strip())
            else:
                # Try to parse as decimal hours
                try:
                    return float(time_str)
                except:
                    return 0
            
            return hours + minutes/60
        except:
            return 0
    
    time_df['time_hours'] = time_df['time_spent'].apply(time_to_hours)
    time_df['total_time_hours'] = time_df['total_time_spent'].apply(time_to_hours)
    
    return time_df, help_ticket_df

def create_executive_summary(ws, time_df, help_ticket_df):
    """Create executive summary with key insights"""
    
    # Title
    ws['A1'] = "🏢 IT HELPDESK EXECUTIVE SUMMARY - AUGUST 2025"
    ws['A1'].font = Font(size=18, bold=True, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='FF2F5597', end_color='FF2F5597', fill_type='solid')
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Key metrics calculation
    total_time_hours = time_df['time_hours'].sum()
    total_tickets = len(help_ticket_df)
    unique_tickets = help_ticket_df['Help Ticket Number'].nunique() if 'Help Ticket Number' in help_ticket_df.columns else total_tickets
    total_technicians = time_df['technician'].nunique()
    avg_time_per_ticket = total_time_hours / max(unique_tickets, 1)
    
    # Most productive technician
    tech_productivity = time_df.groupby('technician').agg({
        'time_hours': 'sum',
        'ticket_number': 'nunique'
    })
    tech_productivity['tickets_per_hour'] = tech_productivity['ticket_number'] / tech_productivity['time_hours']
    most_productive = tech_productivity['tickets_per_hour'].idxmax()
    
    # Key metrics display
    metrics = [
        ("📊 TOTAL HOURS LOGGED", f"{total_time_hours:.1f} hours", "A3", "B3"),
        ("🎫 TOTAL TICKETS", f"{unique_tickets:,} tickets", "D3", "E3"),
        ("👥 ACTIVE TECHNICIANS", f"{total_technicians} people", "G3", "H3"),
        ("⏱️ AVG TIME/TICKET", f"{avg_time_per_ticket:.1f} hours", "A5", "B5"),
        ("🏆 TOP PERFORMER", f"{most_productive}", "D5", "E5"),
        ("📈 PRODUCTIVITY", f"{unique_tickets/total_time_hours:.1f} tickets/hour", "G5", "H5")
    ]
    
    for metric, value, cell1, cell2 in metrics:
        ws[cell1] = metric
        ws[cell1].font = Font(bold=True, size=11)
        ws[cell2] = value
        ws[cell2].font = Font(size=11, color='FF2F5597')
        ws[cell2].fill = PatternFill(start_color='FFE7F3FF', end_color='FFE7F3FF', fill_type='solid')
    
    return 8

def create_technician_analysis(ws, time_df, start_row):
    """Create detailed technician performance analysis"""
    
    ws[f'A{start_row}'] = "👥 TECHNICIAN PERFORMANCE ANALYSIS"
    ws[f'A{start_row}'].font = Font(size=14, bold=True, color='FFFFFF')
    ws[f'A{start_row}'].fill = PatternFill(start_color='FF70AD47', end_color='FF70AD47', fill_type='solid')
    ws.merge_cells(f'A{start_row}:H{start_row}')
    
    # Calculate technician stats
    tech_stats = time_df.groupby('technician').agg({
        'time_hours': ['sum', 'mean', 'count'],
        'ticket_number': 'nunique'
    }).round(2)
    
    tech_stats.columns = ['Total_Hours', 'Avg_Hours_Per_Entry', 'Total_Entries', 'Unique_Tickets']
    tech_stats['Efficiency_Score'] = (tech_stats['Unique_Tickets'] / tech_stats['Total_Hours']).round(2)
    tech_stats['Workload_Score'] = (tech_stats['Total_Hours'] / tech_stats['Total_Hours'].sum() * 100).round(1)
    tech_stats = tech_stats.sort_values('Total_Hours', ascending=False).reset_index()
    
    # Headers
    headers = ['Technician', 'Total Hours', 'Unique Tickets', 'Efficiency (T/H)', 'Workload %', 'Avg Time/Entry', 'Performance Rating']
    header_row = start_row + 2
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Add technician data
    for idx, row in tech_stats.iterrows():
        data_row = header_row + 1 + idx
        
        ws.cell(row=data_row, column=1, value=row['technician'])
        ws.cell(row=data_row, column=2, value=row['Total_Hours'])
        ws.cell(row=data_row, column=3, value=row['Unique_Tickets'])
        ws.cell(row=data_row, column=4, value=row['Efficiency_Score'])
        ws.cell(row=data_row, column=5, value=f"{row['Workload_Score']}%")
        ws.cell(row=data_row, column=6, value=row['Avg_Hours_Per_Entry'])
        
        # Performance rating based on efficiency
        efficiency = row['Efficiency_Score']
        if efficiency >= 2.0:
            rating = "⭐⭐⭐ Excellent"
        elif efficiency >= 1.5:
            rating = "⭐⭐ Good"
        elif efficiency >= 1.0:
            rating = "⭐ Average"
        else:
            rating = "⚠️ Needs Improvement"
        
        ws.cell(row=data_row, column=7, value=rating)
        
        # Color coding rows
        if idx % 2 == 0:
            for col in range(1, 8):
                ws.cell(row=data_row, column=col).fill = PatternFill(start_color='FFF2F2F2', end_color='FFF2F2F2', fill_type='solid')
    
    # Add conditional formatting for efficiency column
    efficiency_range = f"D{header_row+1}:D{header_row+len(tech_stats)}"
    rule = ColorScaleRule(start_type='min', start_color='FFFF6B6B',
                         mid_type='percentile', mid_value=50, mid_color='FFFFEB3B',
                         end_type='max', end_color='FF4ECDC4')
    ws.conditional_formatting.add(efficiency_range, rule)
    
    return start_row + len(tech_stats) + 5

def create_ticket_insights(ws, help_ticket_df, start_row):
    """Create ticket category and status insights"""
    
    ws[f'A{start_row}'] = "🎫 TICKET INSIGHTS & CATEGORIES"
    ws[f'A{start_row}'].font = Font(size=14, bold=True, color='FFFFFF')
    ws[f'A{start_row}'].fill = PatternFill(start_color='FFFFC000', end_color='FFFFC000', fill_type='solid')
    ws.merge_cells(f'A{start_row}:F{start_row}')
    
    # Category analysis
    if 'Category' in help_ticket_df.columns:
        category_counts = help_ticket_df['Category'].value_counts().head(8)
        total_tickets = len(help_ticket_df)
        
        # Headers
        headers = ['Category', 'Count', 'Percentage', 'Priority Level']
        header_row = start_row + 2
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='FFFF6B35', end_color='FFFF6B35', fill_type='solid')
        
        # Add category data
        for idx, (category, count) in enumerate(category_counts.items()):
            data_row = header_row + 1 + idx
            percentage = (count / total_tickets) * 100
            
            ws.cell(row=data_row, column=1, value=str(category))
            ws.cell(row=data_row, column=2, value=count)
            ws.cell(row=data_row, column=3, value=f"{percentage:.1f}%")
            
            # Assign priority based on volume
            if percentage >= 20:
                priority = "🔴 Critical Focus"
            elif percentage >= 10:
                priority = "🟡 High Priority"
            elif percentage >= 5:
                priority = "🟢 Monitor"
            else:
                priority = "⚪ Low Volume"
            
            ws.cell(row=data_row, column=4, value=priority)
    
    # Status analysis
    if 'Status' in help_ticket_df.columns:
        status_counts = help_ticket_df['Status'].value_counts()
        
        ws[f'A{start_row + 15}'] = "📊 TICKET STATUS BREAKDOWN"
        ws[f'A{start_row + 15}'].font = Font(size=12, bold=True)
        
        status_row = start_row + 17
        for idx, (status, count) in enumerate(status_counts.items()):
            ws.cell(row=status_row + idx, column=1, value=str(status))
            ws.cell(row=status_row + idx, column=2, value=count)
            ws.cell(row=status_row + idx, column=3, value=f"{(count/len(help_ticket_df)*100):.1f}%")
    
    return start_row + 25

def create_time_analysis(ws, time_df, start_row):
    """Create time-based analysis and trends"""
    
    ws[f'A{start_row}'] = "⏰ TIME ANALYSIS & TRENDS"
    ws[f'A{start_row}'].font = Font(size=14, bold=True, color='FFFFFF')
    ws[f'A{start_row}'].fill = PatternFill(start_color='FF9B59B6', end_color='FF9B59B6', fill_type='solid')
    ws.merge_cells(f'A{start_row}:F{start_row}')
    
    # Time distribution analysis
    time_ranges = [
        ("Quick (< 0.5h)", time_df[time_df['time_hours'] < 0.5].shape[0]),
        ("Standard (0.5-2h)", time_df[(time_df['time_hours'] >= 0.5) & (time_df['time_hours'] < 2)].shape[0]),
        ("Complex (2-4h)", time_df[(time_df['time_hours'] >= 2) & (time_df['time_hours'] < 4)].shape[0]),
        ("Extended (4h+)", time_df[time_df['time_hours'] >= 4].shape[0])
    ]
    
    headers = ['Time Range', 'Count', 'Percentage', 'Recommendation']
    header_row = start_row + 2
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='FF8E44AD', end_color='FF8E44AD', fill_type='solid')
    
    total_entries = len(time_df)
    recommendations = [
        "✅ Efficient resolution",
        "📋 Standard process",
        "🔍 Review complexity",
        "⚠️ Escalation needed"
    ]
    
    for idx, ((range_name, count), rec) in enumerate(zip(time_ranges, recommendations)):
        data_row = header_row + 1 + idx
        percentage = (count / total_entries) * 100
        
        ws.cell(row=data_row, column=1, value=range_name)
        ws.cell(row=data_row, column=2, value=count)
        ws.cell(row=data_row, column=3, value=f"{percentage:.1f}%")
        ws.cell(row=data_row, column=4, value=rec)
    
    # Peak hours analysis (if date data available)
    if 'date_entered' in time_df.columns:
        try:
            time_df['date_entered'] = pd.to_datetime(time_df['date_entered'], errors='coerce')
            time_df['hour'] = time_df['date_entered'].dt.hour
            
            peak_hours = time_df.groupby('hour')['time_hours'].sum().sort_values(ascending=False).head(3)
            
            ws[f'A{start_row + 8}'] = "🕐 PEAK ACTIVITY HOURS"
            ws[f'A{start_row + 8}'].font = Font(size=12, bold=True)
            
            for idx, (hour, total_time) in enumerate(peak_hours.items()):
                ws.cell(row=start_row + 10 + idx, column=1, value=f"{hour:02d}:00")
                ws.cell(row=start_row + 10 + idx, column=2, value=f"{total_time:.1f}h")
        except:
            pass
    
    return start_row + 15

def create_recommendations(ws, time_df, help_ticket_df, start_row):
    """Create actionable recommendations based on data analysis"""
    
    ws[f'A{start_row}'] = "💡 ACTIONABLE RECOMMENDATIONS"
    ws[f'A{start_row}'].font = Font(size=14, bold=True, color='FFFFFF')
    ws[f'A{start_row}'].fill = PatternFill(start_color='FF27AE60', end_color='FF27AE60', fill_type='solid')
    ws.merge_cells(f'A{start_row}:H{start_row}')
    
    # Calculate insights for recommendations
    total_time = time_df['time_hours'].sum()
    avg_time_per_ticket = total_time / help_ticket_df['Help Ticket Number'].nunique() if 'Help Ticket Number' in help_ticket_df.columns else 1
    
    # Top time-consuming categories
    if 'Category' in help_ticket_df.columns:
        top_categories = help_ticket_df['Category'].value_counts().head(3).index.tolist()
    else:
        top_categories = ["Hardware", "Software", "Network"]
    
    recommendations = [
        ("🎯 Focus Areas", f"Prioritize automation for: {', '.join(top_categories[:2])}", "High Impact"),
        ("⚡ Efficiency", f"Current avg: {avg_time_per_ticket:.1f}h/ticket. Target: <2h", "Process Improvement"),
        ("📚 Knowledge Base", "Create self-service guides for top 5 issues", "Reduce Volume"),
        ("👥 Training", "Cross-train technicians on high-volume categories", "Resource Optimization"),
        ("📊 Monitoring", "Implement real-time SLA tracking dashboard", "Performance Management"),
        ("🔄 Process", "Review tickets >4 hours for process improvements", "Quality Assurance")
    ]
    
    headers = ['Category', 'Recommendation', 'Impact Level', 'Priority']
    header_row = start_row + 2
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='FF229954', end_color='FF229954', fill_type='solid')
    
    priorities = ["🔴 Immediate", "🟡 This Month", "🟢 Next Quarter", "🔵 Long-term", "🟣 Ongoing", "⚪ As Needed"]
    
    for idx, (category, recommendation, impact) in enumerate(recommendations):
        data_row = header_row + 1 + idx
        
        ws.cell(row=data_row, column=1, value=category)
        ws.cell(row=data_row, column=2, value=recommendation)
        ws.cell(row=data_row, column=3, value=impact)
        ws.cell(row=data_row, column=4, value=priorities[idx])
        
        # Alternate row colors
        if idx % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=data_row, column=col).fill = PatternFill(start_color='FFE8F5E8', end_color='FFE8F5E8', fill_type='solid')
    
    return start_row + len(recommendations) + 5

def format_worksheet(ws):
    """Apply final formatting to the worksheet"""
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add borders to all cells with content
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border

def main():
    """Main function to create comprehensive enhanced report"""
    
    print("🚀 Creating Comprehensive IT Helpdesk Report Enhancement...")
    print("=" * 60)
    
    try:
        # Load and analyze data
        time_df, help_ticket_df = analyze_and_enhance_data()
        
        # Load enhanced workbook
        wb = load_workbook(enhanced_file)
        
        # Create comprehensive analytics sheet
        if 'Comprehensive_Analytics' in wb.sheetnames:
            wb.remove(wb['Comprehensive_Analytics'])
        
        ws = wb.create_sheet('Comprehensive_Analytics', 0)  # Make it the first sheet
        
        # Build the comprehensive report
        current_row = 1
        current_row = create_executive_summary(ws, time_df, help_ticket_df) + 2
        current_row = create_technician_analysis(ws, time_df, current_row) + 2
        current_row = create_ticket_insights(ws, help_ticket_df, current_row) + 2
        current_row = create_time_analysis(ws, time_df, current_row) + 2
        current_row = create_recommendations(ws, time_df, help_ticket_df, current_row)
        
        # Apply final formatting
        format_worksheet(ws)
        
        # Save the enhanced workbook
        wb.save(enhanced_file)
        
        # Print summary
        print("✅ ENHANCEMENT COMPLETE!")
        print("=" * 60)
        print(f"📁 Enhanced file: Monthly_Reporting_Aug_Enhanced.xlsx")
        print(f"📊 Total time analyzed: {time_df['time_hours'].sum():.1f} hours")
        print(f"🎫 Total tickets: {len(help_ticket_df):,}")
        print(f"👥 Technicians: {time_df['technician'].nunique()}")
        print(f"⚡ Average time per ticket: {time_df['time_hours'].sum() / help_ticket_df['Help Ticket Number'].nunique():.1f} hours")
        
        print("\n🎯 NEW FEATURES ADDED:")
        print("• Executive summary with key metrics")
        print("• Technician performance analysis with ratings")
        print("• Ticket category insights and priorities")
        print("• Time distribution analysis")
        print("• Actionable recommendations")
        print("• Professional formatting and color coding")
        print("• Conditional formatting for data visualization")
        print("• Industry-standard KPI tracking")
        
        print("\n📈 INDUSTRY-STANDARD METRICS INCLUDED:")
        print("• Efficiency scores (tickets per hour)")
        print("• Workload distribution analysis")
        print("• Category prioritization matrix")
        print("• Time complexity breakdown")
        print("• Performance ratings and recommendations")
        
        # Show top insights
        top_performer = time_df.groupby('technician')['time_hours'].sum().idxmax()
        top_hours = time_df.groupby('technician')['time_hours'].sum().max()
        
        print(f"\n🏆 KEY INSIGHTS:")
        print(f"• Top performer: {top_performer} ({top_hours:.1f} hours)")
        print(f"• Most common category: {help_ticket_df['Category'].mode().iloc[0] if 'Category' in help_ticket_df.columns else 'N/A'}")
        print(f"• Productivity rate: {len(help_ticket_df) / time_df['time_hours'].sum():.1f} tickets/hour")
        
    except Exception as e:
        print(f"❌ Error creating enhanced report: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
